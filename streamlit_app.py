import streamlit as st
import pandas as pd
import docx
import os
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_community.vectorstores import FAISS
from langchain.prompts import PromptTemplate
from langchain.chains import RetrievalQA
from langchain.text_splitter import RecursiveCharacterTextSplitter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import smtplib
from email.message import EmailMessage

# Verify API key
openai_api_key = os.getenv("OPENAI_API_KEY")
if not openai_api_key:
    st.error("OpenAI API key not found! Set it in your environment.")
    st.stop()

# Verify required files exist
required_files = ["sas_code_example.docx", "campaign_requirements_example.docx",
                  "project_details.txt", "segment_details.txt", "syf logo.png"]
missing_files = [file for file in required_files if not os.path.exists(file)]
if missing_files:
    st.error(f"Missing required file(s): {', '.join(missing_files)}")
    st.stop()

# UI Elements
st.image("syf logo.png", width=100)
st.title("AI-Based Campaign Operation Programming Solution (AI-COPS)")

# Load and process text files
project_details = pd.read_csv("project_details.txt", sep='|')
segment_details = pd.read_csv("segment_details.txt", sep='|')

# Load docx files function
def load_docx(file_path):
    doc = docx.Document(file_path)
    return '\n'.join(para.text for para in doc.paragraphs)

sas_text = load_docx("sas_code_example.docx")
campaign_text = load_docx("campaign_requirements_example.docx")

# Vector Database setup
from langchain.text_splitter import RecursiveCharacterTextSplitter
text_splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
all_chunks = text_splitter.split_text(sas_text) + text_splitter.split_text(campaign_text)

embeddings = OpenAIEmbeddings(openai_api_key=openai_api_key)
vector_db = FAISS.from_texts(all_chunks, embeddings)

# UI Input
wf_number = st.text_input("🔢 Enter Workfront Number (Numeric Only)")
user_email = st.text_input("📧 Enter Your Email")

# Main Logic
if st.button("🚀 Submit"):
    if not wf_number.isdigit():
        st.error("Workfront number must be numeric.")
    else:
        wf_number_int = int(wf_number)
        project_info = project_details[project_details['WFNO'] == wf_number_int]
        segment_info = segment_details[segment_details['WFNO'] == wf_number_int]

        if project_info.empty or segment_info.empty:
            st.error("No matching details found for this Workfront number.")
        else:
            # Extracting required fields safely
            campaign_req = project_info.iloc[0]['Campaign Requirements']
            
            suppress_fields = ['Marketing', 'Risk', 'Optout']
            suppressions = [
                field for field in suppress_fields
                if field in project_info.columns and project_info.iloc[0][field] == 'Y'
            ]

            outfile_type = project_info.iloc[0]['Outfile Required'] if 'Outfile Required' in project_info.columns else 'Not Specified'
            misc_info = project_info.iloc[0]['Campaign Name'] if 'Campaign Name' in project_info.columns else 'None'

            standard_prompt = (
                f"Write SAS code for a campaign to target '{project_info.iloc[0]['Campaign Requirements']}' "
                f"with suppressions: {', '.join(suppressions)}. "
                f"Outfile type: {outfile_type}. Misc info: {misc_info if 'misc_info' in locals() else 'None'}."
            )

            llm = ChatOpenAI(model_name="gpt-4o-mini", temperature=0, openai_api_key=openai_api_key)
            prompt_template = PromptTemplate(
                input_variables=["context", "question"],
                template="Context: {context}\n\nTask: {question}",
            )

            vector_db = FAISS.from_texts(all_chunks, embeddings)
            qa_chain = RetrievalQA.from_chain_type(
                llm=llm,
                chain_type="stuff",
                retriever=vector_db.as_retriever(),
                chain_type_kwargs={"prompt": prompt_template},
            )

            sas_code_response = qa_chain.run(standard_prompt)
            st.subheader("📄 Generated SAS Code")
            st.code(sas_code_response, language='sas')

            # Excel workbook creation
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Project Details"
            for row in dataframe_to_rows(project_info, index=False, header=True):
                ws1.append(row)

            wb.create_sheet("Waterfall").append(["Waterfall Data Placeholder"])
            ws3 = wb.create_sheet("Segment Details")
            for row in dataframe_to_rows(segment_info, index=False, header=True):
                ws3.append(row)
            wb.create_sheet("Output File Layout").append(["Output File Layout", outfile_type])

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            

            # Email Logic (unchanged)
            try:
                EMAIL_ADDRESS = "mirza.22sept@gmail.com"
                EMAIL_APP_PASSWORD = os.getenv("EMAIL_APP_PASSWORD")
                if not EMAIL_APP_PASSWORD:
                    st.error("Email App Password not set in environment variables.")
                    st.stop()
                msg = EmailMessage()
                msg['Subject'] = f"Campaign Details - Workfront {wf_number}"
                msg['From'] = EMAIL_APP_PASSWORD
                msg['To'] = user_email
                msg.set_content(f"Attached campaign details for Workfront Number {wf_number}.")
                msg.add_attachment(excel_buffer.read(), maintype='application',
                                   subtype='xlsx', filename=f"Campaign_{wf_number}.xlsx")

                with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                    smtp.starttls()
                    smtp.login(EMAIL_ADDRESS, EMAIL_APP_PASSWORD)
                    smtp.send_message(msg)

                st.success(f"Excel file successfully sent to {user_email}")
            except Exception as e:
                st.error(f"Email sending failed: {e}")
